from selenium import webdriver
import time
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import configparser

def consultprofiles(my_login, my_password, file_xlsx):
    dataframe = openpyxl.load_workbook(file_xlsx)
    dataframe1 = dataframe.active

    driver = webdriver.Chrome()

    nb_profils_consultes = 0
    wait = WebDriverWait(driver, 10)
    x = False
    
    print(f"Ouverture de la fenêtre ...")
    print(f"Veuillez à ne pas fermer la fenêtre avant la fin du traitement !")

    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            url = f"{col[row].value}"
            driver.get(url)
            if not x:
                driver.find_element(By.CLASS_NAME, "cta.bg-blue.header__container__button_block__item").click()
                login = wait.until(EC.visibility_of_element_located((By.CLASS_NAME, "LoginModalContainer__SInputForm-sc-1fx857l-3")))
                login.send_keys(my_login)
                password = driver.find_element(By.NAME, "password")
                password.send_keys(my_password)
                driver.find_element(By.CLASS_NAME, "dBlock").click()
                if x:
                    driver.find_element(By.CLASS_NAME, "dInlineB.pad10_10.vaM.theme-notif-menu-crush").click()
                x = True
            time.sleep(10)
            nb_profils_consultes += 1 
            total_profils = len(col)
            print(f"Profils consultés : {nb_profils_consultes} / {total_profils}")
    print(f"Profils consultés : {nb_profils_consultes} / {total_profils}")
    driver.quit()  

def read_credentials_from_config(file_path):
    config = configparser.ConfigParser()
    config.read(file_path)
    return config['Credentials']

def main():
    default_config = read_credentials_from_config('config.ini')
    local_config = read_credentials_from_config('config.local.ini')

    username = local_config.get('username', default_config['username'])
    password = local_config.get('password', default_config['password'])
    file_name = local_config.get('file_name', default_config['file_name'])

    consultprofiles(username, password, file_name)

if __name__ == "__main__":
    main()
